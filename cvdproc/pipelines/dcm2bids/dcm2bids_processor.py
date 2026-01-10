import subprocess
import os
import shutil
import nibabel as nib
import numpy as np
import json
import pandas as pd
import glob
import pydicom
import re
from collections import defaultdict
from bids.cli import layout
from openpyxl import Workbook
from openpyxl.styles import Font
from bids.layout import BIDSLayout
from requests import session
from openpyxl.utils import get_column_letter

import nipype
from cvdproc.pipelines.smri.fsl.deface_nipype import FSLDeface

class Dcm2BidsProcessor:
    def __init__(self, BIDS_root_folder):
        self.BIDS_root_folder = BIDS_root_folder

    def initialize(self):
        # check if the output directory exists
        if not os.path.exists(self.BIDS_root_folder):
            print('Creating output directory...')
            os.makedirs(self.BIDS_root_folder)
        else:
            print('Output directory already exists.')

        subprocess.run(['dcm2bids_scaffold', '-o', self.BIDS_root_folder])

        # Experimental: create 'population', 'workflows' folder in derivatives folder
        derivatives_folder = os.path.join(self.BIDS_root_folder, 'derivatives')
        population_folder = os.path.join(derivatives_folder, 'population')
        workflows_folder = os.path.join(derivatives_folder, 'workflows')
        os.makedirs(population_folder, exist_ok=True)
        os.makedirs(workflows_folder, exist_ok=True)

        # === Overwrite participants.tsv ===
        participants_tsv = os.path.join(self.BIDS_root_folder, "participants.tsv")
        col_order = ["participant_id", "session_id", "name", "imaging_id", "acq_time", "institution_name", "age", "sex", "convert_time"]

        with open(participants_tsv, "w") as f:
            f.write("\t".join(col_order) + "\n")

        print("participants.tsv created/reset.")

        # === Overwrite participants.json ===
        participants_json = os.path.join(self.BIDS_root_folder, "participants.json")
        participants_metadata = {
            "participant_id": {
                "Description": "Unique participant identifier following BIDS convention (sub-XX)"
            },
            "session_id": {
                "Description": "Session identifier following BIDS convention (ses-XX)"
            },
            "name": {
                "Description": "Patient name extracted from DICOM header"
            },
            "imaging_id": {
                "Description": "Patient ID extracted from DICOM header"
            },
            "acq_time": {
                "Description": "Acquisition date extracted from DICOM StudyDate"
            },
            "institution_name": {
                "Description": "Name of the institution where the imaging was acquired"
            },
            "age": {
                "Description": "Patient age from DICOM header"
            },
            "sex": {
                "Description": "Patient sex from DICOM header"
            },
            "convert_time": {
                "Description": "Current time"
            }
        }

        with open(participants_json, "w") as f:
            json.dump(participants_metadata, f, indent=4)

        print("participants.json created/reset.")

        print('Initialization completed.')

    def _build_ignore_predicate(self, ignore_list):
        """
        Compile ignore patterns as case-insensitive regexes.
        Returns a predicate that tests a SeriesDescription string.
        """
        if not ignore_list:
            return lambda _: False
        pats = []
        for pat in ignore_list:
            # Treat plain strings as substring regex; allow full regex as-is
            # Escape only if string has no regex metacharacters.
            if re.search(r"[.^$*+?{}\[\]|()\\]", pat):
                pats.append(re.compile(pat, re.IGNORECASE))
            else:
                pats.append(re.compile(re.escape(pat), re.IGNORECASE))
        def _pred(desc):
            if desc is None:
                return False
            txt = str(desc)
            return any(p.search(txt) for p in pats)
        return _pred

    def _link_or_copy(self, src, dst):
        """
        Try to create a hard link for speed; fall back to copy if cross-device.
        """
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        try:
            os.link(src, dst)
        except Exception:
            shutil.copy2(src, dst)

    def filter_dicom_by_series_description(self, src_dir, out_dir, ignore_patterns):
        """
        Build a filtered dicom tree under out_dir by excluding series whose
        SeriesDescription matches any ignore pattern.
        Returns (kept_count, ignored_count).
        """
        pred = self._build_ignore_predicate(ignore_patterns)
        series_files = defaultdict(list)
        series_desc  = {}

        for root, _, files in os.walk(src_dir):
            for f in files:
                fpath = os.path.join(root, f)
                try:
                    ds = pydicom.dcmread(
                        fpath, stop_before_pixels=True,
                        specific_tags=["SeriesInstanceUID", "SeriesDescription"]
                    )
                except Exception:
                    continue
                uid = getattr(ds, "SeriesInstanceUID", None)
                if uid is None:
                    continue
                series_files[uid].append(fpath)
                if uid not in series_desc:
                    series_desc[uid] = str(getattr(ds, "SeriesDescription", ""))

        kept_uids = [uid for uid, desc in series_desc.items() if not pred(desc)]
        ignored_uids = [uid for uid in series_desc.keys() if uid not in kept_uids]

        kept_count = 0
        ignored_count = 0

        # Write per-series folders for clarity and safety
        def _sanitize(s):
            return re.sub(r"[^A-Za-z0-9._-]+", "_", s)[:80] or "NA"

        for uid in kept_uids:
            desc = _sanitize(series_desc.get(uid, ""))
            series_out = os.path.join(out_dir, f"series-{desc}_{uid}")
            for src in series_files[uid]:
                dst = os.path.join(series_out, os.path.basename(src))
                self._link_or_copy(src, dst)
                kept_count += 1

        for uid in ignored_uids:
            ignored_count += len(series_files[uid])

        return kept_count, ignored_count

    def convert(self, config_file, dicom_directory, subject_id, session_id,
                ignore_patterns=None, keep_temp=False):
        """
        Optionally pre-filter DICOMs by SeriesDescription before running dcm2bids.
        """
        # Decide input directory (original or filtered temp)
        use_dir = dicom_directory
        tmp_root = os.path.join(self.BIDS_root_folder, "tmp_dcm2bids")
        if ignore_patterns:
            # Ensure tmp root exists
            os.makedirs(tmp_root, exist_ok=True)
            tag = f"sub-{subject_id}" + (f"_ses-{session_id}" if session_id else "")
            tmp_dicom = os.path.join(tmp_root, f"{tag}_dicomtemp")
            if os.path.exists(tmp_dicom):
                shutil.rmtree(tmp_dicom)
            os.makedirs(tmp_dicom, exist_ok=True)

            kept, ignored = self.filter_dicom_by_series_description(
                src_dir=dicom_directory,
                out_dir=tmp_dicom,
                ignore_patterns=ignore_patterns
            )
            print(f"[DICOM filter] Kept: {kept} files; Ignored: {ignored} files.")
            use_dir = tmp_dicom

        subprocess.run([
            'dcm2bids',
            '-d', use_dir,
            '-p', subject_id,
            '-s', session_id,
            '-c', config_file,
            '-o', self.BIDS_root_folder,
            '--auto_extract_entities'
        ], check=False)

        # Cleanup temporary filtered dicom dir if not needed
        if ignore_patterns and not keep_temp:
            try:
                shutil.rmtree(use_dir)
            except Exception:
                pass

        # === cropped 3D replacement block ===
        tmp_dcm2bids_folder = os.path.join(self.BIDS_root_folder, 'tmp_dcm2bids')
        if os.path.exists(tmp_dcm2bids_folder):
            print('Checking for cropped 3D images...')
            cropped_image_count = 0
            if session_id:
                subject_temp_folder = os.path.join(tmp_dcm2bids_folder, f'sub-{subject_id}_ses-{session_id}')
                subject_bids_folder = os.path.join(self.BIDS_root_folder, f'sub-{subject_id}', f'ses-{session_id}')
            else:
                subject_temp_folder = os.path.join(tmp_dcm2bids_folder, f'sub-{subject_id}')
                subject_bids_folder = os.path.join(self.BIDS_root_folder, f'sub-{subject_id}')

            for file in os.listdir(subject_temp_folder):
                if 'Crop' in file and '.nii' in file:
                    original_file = file.replace('_Crop_1', '')
                    original_file_path = os.path.join(subject_temp_folder, original_file)
                    if not os.path.exists(original_file_path):
                        print('Found cropped 3D image, moving to BIDS folder...')
                        cropped_image = os.path.join(subject_temp_folder, file)
                        cropped_nii = nib.load(cropped_image)
                        cropped_shape = cropped_nii.shape
                        cropped_voxel_size = cropped_nii.header.get_zooms()
                        
                        files_matched = 0
                        for root, dirs, files in os.walk(subject_bids_folder):
                            for file in files:
                                if '.nii' in file:
                                    nii_file = os.path.join(root, file)
                                    nii = nib.load(nii_file)
                                    if (nii.shape[0] == cropped_shape[0] and
                                        nii.shape[1] == cropped_shape[1] and
                                        nii.header.get_zooms() == cropped_voxel_size):
                                        bids_converted_file_path = nii_file
                                        files_matched += 1
                                        cropped_image_count += 1

                        if files_matched == 1:
                            print(f'{cropped_image} -> {bids_converted_file_path}')
                            os.rename(cropped_image, bids_converted_file_path)
                        elif files_matched > 1:
                            print('More than one file matched, please check the files manually.')
                        else:
                            print('No file matched, please check the files manually.')

            if cropped_image_count == 0:
                print('No cropped 3D image found.')
        else:
            print('No temporary dcm2bids folder found, or it is not in the BIDS root folder.')

    
    def fix_intendedfor_for_subject_session(self, subject_id, session_id):
        """
        Fix 'IntendedFor' fields only under a specific subject/session.
        """
        target_dir = os.path.join(self.BIDS_root_folder, f"sub-{subject_id}", f"ses-{session_id}")
        if not os.path.exists(target_dir):
            print(f"[WARN] Target directory not found: {target_dir}")
            return

        fixed_count = 0
        for root, _, files in os.walk(target_dir):
            for file in files:
                if file.endswith(".json"):
                    json_path = os.path.join(root, file)
                    try:
                        with open(json_path, "r", encoding="utf-8") as f:
                            data = json.load(f)

                        if "IntendedFor" in data:
                            original = data["IntendedFor"]
                            modified = None

                            if isinstance(original, str):
                                if not original.startswith("ses-") and "ses-" in original:
                                    modified = "ses-" + original.split("ses-", 1)[1]
                            elif isinstance(original, list):
                                changed = False
                                new_list = []
                                for item in original:
                                    if not item.startswith("ses-") and "ses-" in item:
                                        new_list.append("ses-" + item.split("ses-", 1)[1])
                                        changed = True
                                    else:
                                        new_list.append(item)
                                if changed:
                                    modified = new_list

                            if modified is not None:
                                data["IntendedFor"] = modified
                                with open(json_path, "w", encoding="utf-8") as f:
                                    json.dump(data, f, indent=4)
                                fixed_count += 1
                                print(f"[FIXED] IntendedFor: {json_path}")

                    except Exception as e:
                        print(f"[ERROR] Failed to process {json_path}: {e}")
                        continue

        if fixed_count > 0:
            print(f"Fixed {fixed_count} JSON file(s) for sub-{subject_id} ses-{session_id}")
    
    def fix_dwi_bvec_bval(self, subject_id, session_id, fix_config):
        dwi_dir = os.path.join(
            self.BIDS_root_folder,
            f"sub-{subject_id}",
            f"ses-{session_id}",
            "dwi"
        )

        if not os.path.exists(dwi_dir):
            print(f"[WARN] No DWI directory found for sub-{subject_id}, ses-{session_id}. Skipping fix.")
            return

        for file in os.listdir(dwi_dir):
            for fix_item in fix_config:
                match_str = fix_item.get("match", "")
                if match_str in file and file.endswith(".nii.gz"):
                    base = file.replace(".nii.gz", "")
                    bvec_path = os.path.join(dwi_dir, base + ".bvec")
                    bval_path = os.path.join(dwi_dir, base + ".bval")

                    if os.path.exists(bvec_path) and os.path.exists(fix_item["bvec"]):
                        print(f"Replacing {bvec_path} with {fix_item['bvec']}")
                        shutil.copyfile(fix_item["bvec"], bvec_path)

                    if os.path.exists(bval_path) and os.path.exists(fix_item["bval"]):
                        print(f"Replacing {bval_path} with {fix_item['bval']}")
                        shutil.copyfile(fix_item["bval"], bval_path)
    
    def fix_perf_aslcontext(self, subject_id, session_id, fix_config):
        asl_dir = os.path.join(
            self.BIDS_root_folder,
            f"sub-{subject_id}",
            f"ses-{session_id}",
            "perf"
        )

        if not os.path.exists(asl_dir):
            print(f"[WARN] No perf directory found for sub-{subject_id}, ses-{session_id}. Skipping fix.")
            return

        for file in os.listdir(asl_dir):
            for fix_item in fix_config:
                match_str = fix_item.get("match", "")
                if match_str in file and file.endswith(".nii.gz"):
                    base = file.replace("asl.nii.gz", "")
                    aslcontext_path = os.path.join(asl_dir, base + "aslcontext.tsv")
                    if os.path.exists(fix_item["aslcontext"]):
                        print(f"Replacing or creating {aslcontext_path} with {fix_item['aslcontext']}")
                        shutil.copyfile(fix_item["aslcontext"], aslcontext_path)
    
    def deface_anat(self, subject_id, session_id, suffix_list=['T1w', 'T2w', 'FLAIR']):
        """
        Deface anatomical images for a specific subject and session.
        """
        anat_dir = os.path.join(
            self.BIDS_root_folder,
            f"sub-{subject_id}",
            f"ses-{session_id}",
            "anat"
        )
        # Deface each file ending with <suffix>.nii.gz and overwrite the original.
        if not os.path.exists(anat_dir):
            print(f"[WARN] No anat directory found for sub-{subject_id}, ses-{session_id}. Skipping deface.")
            return
        for file in os.listdir(anat_dir):
            for suffix in suffix_list:
                if file.endswith(f"{suffix}.nii.gz"):
                    file_path = os.path.join(anat_dir, file)
                    print(f"Defacing {file_path}...")
                    deface = FSLDeface()
                    deface.inputs.input_image = file_path
                    deface.inputs.output_image = file_path  # overwrite the original file
                    result = deface.run()
                    if result:
                        print(f"Defaced {file_path} successfully.")
                    else:
                        print(f"[ERROR] Failed to deface {file_path}.")
    
    def find_first_dicom(self, dicom_root):
        """Recursively find the first dicom file under dicom_root"""
        for root, dirs, files in os.walk(dicom_root):
            for f in files:
                fpath = os.path.join(root, f)
                try:
                    ds = pydicom.dcmread(fpath, stop_before_pixels=True)
                    return ds
                except Exception:
                    continue
        return None

    def update_participants_tsv(self, bids_dir, subject_id, session_id, ds):
        participants_tsv = os.path.join(bids_dir, "participants.tsv")

        col_order = ["participant_id", "session_id", "name", "imaging_id", "acq_time", "institution_name", "age", "sex", "convert_time"]

        if not os.path.exists(participants_tsv):
            with open(participants_tsv, "w") as f:
                f.write("\t".join(col_order) + "\n")

        df = pd.read_csv(participants_tsv, sep="\t")

        name = getattr(ds, "PatientName", "")
        pid = getattr(ds, "PatientID", "")
        study_date = getattr(ds, "StudyDate", "")
        institution_name = getattr(ds, "InstitutionName", "")
        age = getattr(ds, "PatientAge", "")
        sex = getattr(ds, "PatientSex", "")
        # get the current time (system time)
        from datetime import datetime
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        new_row = {
            "participant_id": f"sub-{subject_id}",
            "session_id": f"ses-{session_id}",
            "name": str(name),
            "imaging_id": str(pid),
            "acq_time": str(study_date),
            "institution_name": str(institution_name),
            "age": str(age),
            "sex": str(sex),
            "convert_time": str(current_time)
        }

        df = df[~((df["participant_id"] == new_row["participant_id"]) & (df["session_id"] == new_row["session_id"]))]

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

        df = df[col_order]
        df.to_csv(participants_tsv, sep="\t", index=False)

        print(f"participants.tsv updated for {new_row['participant_id']} {new_row['session_id']}")

    def check_data(self, check_data_config):
        """
        Check BIDS rawdata and/or derivatives files based on filename or json criteria, 
        and save results to check_data.xlsx
        """

        print("Checking BIDS rawdata and derivatives...")

        results = []
        column_titles = ["Subject_id", "Session_id"]

        for check in check_data_config:
            custom_name = check["custom_name"]
            column_titles.extend([custom_name, f"{custom_name}_number"])

        # Collect all subjects (only those present in rawdata)
        all_subjects = sorted([
            d.replace("sub-", "")
            for d in os.listdir(self.BIDS_root_folder)
            if d.startswith("sub-") and os.path.isdir(os.path.join(self.BIDS_root_folder, d))
        ])

        for subject in all_subjects:
            subject_dir = os.path.join(self.BIDS_root_folder, f"sub-{subject}")
            sessions = sorted([
                d.replace("ses-", "")
                for d in os.listdir(subject_dir)
                if d.startswith("ses-") and os.path.isdir(os.path.join(subject_dir, d))
            ])

            for session in sessions:
                row = [subject, session] + [0] * (len(check_data_config) * 2)

                for check_idx, check in enumerate(check_data_config):
                    custom_name = check["custom_name"]
                    criteria = check["criteria"]
                    method = check["method"]
                    derivatives_name = check.get("derivatives_name", None)
                    suffix = check.get("suffix", None)

                    # ---------------- Locate candidate files ----------------
                    filepaths = []
                    if derivatives_name is None:
                        # rawdata
                        session_dir = os.path.join(self.BIDS_root_folder, f"sub-{subject}", f"ses-{session}")
                        if os.path.exists(session_dir):
                            for root, _, files in os.walk(session_dir):
                                for f in files:
                                    if f.endswith(".nii.gz") and (suffix is None or suffix in f):
                                        filepaths.append(os.path.join(root, f))
                    else:
                        # derivatives
                        derivatives_dir = os.path.join(
                            self.BIDS_root_folder, "derivatives", derivatives_name,
                            f"sub-{subject}", f"ses-{session}"
                        )
                        # add files to filepaths, but no strictly to only .nii.gz. And not use suffix
                        if os.path.exists(derivatives_dir):
                            for root, _, files in os.walk(derivatives_dir):
                                for f in files:
                                        filepaths.append(os.path.join(root, f))

                    # ---------------- Matching logic ----------------
                    match_found = False
                    match_count = 0

                    for file in filepaths:
                        if method == "filename":
                            if criteria in os.path.basename(file):
                                match_found = True
                                match_count += 1

                        elif method == "json" and derivatives_name is None:
                            json_file_path = file.replace(".nii.gz", ".json")
                            if os.path.exists(json_file_path):
                                with open(json_file_path, "r") as json_file:
                                    json_data = json.load(json_file)
                                    if all(json_data.get(k) == v for k, v in criteria.items()):
                                        match_found = True
                                        match_count += 1

                    # Update results
                    custom_name_idx = 2 + check_idx * 2
                    custom_name_number_idx = custom_name_idx + 1
                    row[custom_name_idx] = 1 if match_found else 0
                    row[custom_name_number_idx] = match_count

                results.append(row)

        # Save Excel output
        output_dir = os.path.join(self.BIDS_root_folder, "derivatives", "population")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "check_data.xlsx")

        df = pd.DataFrame(results, columns=column_titles)
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Check Data")
            sheet = writer.book["Check Data"]

            # Configure font and column widths
            calibri_font = Font(name="Calibri")
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = calibri_font

            for col_idx, col in enumerate(sheet.columns, start=1):
                max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                sheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        print(f"Results saved to {output_file}")
