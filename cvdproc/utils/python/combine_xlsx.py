import pandas as pd
from openpyxl import load_workbook
import os

def merge_and_save_excel(base_file_path, additional_file_path, match_column_base, match_column_additional, selected_columns, prefix="", output_file_path=None):
    """
    Merges two Excel files based on matching columns and a specified list of columns from the additional file.
    Adds a prefix to the column names from the additional file and saves the result in a new Excel file,
    while preserving the original formatting.

    :param base_file_path: Path to the base Excel file.
    :param additional_file_path: Path to the Excel file containing additional data.
    :param match_column_base: List of column names in the base file used for matching.
    :param match_column_additional: List of column names in the additional file used for matching.
    :param selected_columns: List of column names from the additional file to be added,
                              or the string 'ALL_EXCEPT_MATCH' to include all columns except matching ones.
    :param prefix: Prefix to be added to the column names from the additional file.
    :param output_file_path: The path to save the output Excel file. If None, the output file path will be
                              the base file name with "_merged" appended.
    """
    # Load the base and additional data
    base_df = pd.read_excel(base_file_path)
    additional_df = pd.read_excel(additional_file_path)

    # Convert matching columns to string type to prevent dtype mismatch
    for col in match_column_base:
        base_df[col] = base_df[col].astype(str)
    for col in match_column_additional:
        additional_df[col] = additional_df[col].astype(str)

    # Ensure that match_column_base and match_column_additional have the same number of columns
    if len(match_column_base) != len(match_column_additional):
        raise ValueError("The number of columns in 'match_column_base' must match the number of columns in 'match_column_additional'")

    # Automatically select all columns except matching ones if requested
    if selected_columns == 'ALL_EXCEPT_MATCH':
        selected_columns = [col for col in additional_df.columns if col not in match_column_additional]

    # Filter the selected columns from the additional file
    columns_to_merge = match_column_additional + selected_columns
    additional_df_filtered = additional_df[columns_to_merge]

    # Merge the dataframes on the specified matching columns
    merged_df = pd.merge(base_df, additional_df_filtered, left_on=match_column_base, right_on=match_column_additional, how='left')

    # Add prefix to the selected column names from the additional file
    for col in selected_columns:
        if col in merged_df.columns:
            merged_df.rename(columns={col: prefix + col}, inplace=True)

    # Drop the matching columns from the additional file to avoid duplication
    merged_df.drop(columns=match_column_additional, inplace=True)

    # If no output file path is provided, generate one by appending '_merged' to the base file name
    if output_file_path is None:
        base_name, ext = os.path.splitext(base_file_path)
        output_file_path = f"{base_name}_merged{ext}"

    # Load the original workbook using openpyxl to preserve formatting
    wb = load_workbook(base_file_path)
    sheet = wb.active

    # Write the header (column names) to the first row
    for col_idx, col_name in enumerate(merged_df.columns, 1):
        sheet.cell(row=1, column=col_idx, value=col_name)

    # Write the merged dataframe into the sheet (starting from the second row)
    for r_idx, row in merged_df.iterrows():
        for c_idx, value in enumerate(row, 1):
            current_cell = sheet.cell(row=r_idx + 2, column=c_idx)
            if current_cell.value is None:
                current_cell.value = value

    # Save the workbook with formatting preserved
    wb.save(output_file_path)
    print(f"File saved successfully to {output_file_path}")

# Example usage
if __name__ == "__main__":
    base_file_path = r'D:\WYJ\Research_group\paper\lesion_connection\data\population\data_to_analysis_merged.xlsx'
    additional_file_path = r'D:\WYJ\Research_group\paper\lesion_connection\data\population\fdt\surface_parameters_results_modified.xlsx'
    match_column_base = ['Subject_id', 'Session_id']
    match_column_additional = ['subject', 'session']
    selected_columns = 'ALL_EXCEPT_MATCH'  # Automatically include all columns except matching ones
    prefix = ''

    merge_and_save_excel(base_file_path, additional_file_path, match_column_base, match_column_additional, selected_columns, prefix)